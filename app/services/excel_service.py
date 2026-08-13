"""Xuat Excel bang openpyxl.

Tien va so luong ghi xuong o dang SO voi number_format, khong phai chuoi
'1.500.000 đ' — nho vay nguoi nhan file van SUM/loc duoc trong Excel.
"""
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import EXPORTS_DIR
from app.services import report_service
from app.utils.formatters import order_status_label

MONEY_FMT = '#,##0" đ"'
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F6FED")


def _write_header(sheet, headers: list[str], widths: list[int]) -> None:
    for col, (title, width) in enumerate(zip(headers, widths), start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.freeze_panes = "A2"


def _save(workbook: Workbook, prefix: str):
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    path = EXPORTS_DIR / f"{prefix}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    workbook.save(path)
    return path


def export_orders(orders: list[dict]):
    """Xuat danh sach hoa don (dung danh sach dang hien tren man hinh,
    ton trong bo loc nguoi dung dat). Tra ve duong dan file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hóa đơn"

    _write_header(
        ws,
        ["Mã đơn", "Thời gian", "Khách hàng", "Điện thoại", "Số SP",
         "Tạm tính", "Giảm giá", "Tổng tiền", "Đã hoàn", "Trạng thái"],
        [16, 17, 24, 14, 8, 14, 13, 14, 13, 12],
    )

    for row, order in enumerate(orders, start=2):
        refunded = order.get("refunded", 0)
        status = order_status_label(order)
        ws.cell(row=row, column=1, value=order["order_code"])
        time_cell = ws.cell(row=row, column=2, value=order["created_at"])
        time_cell.number_format = "dd/mm/yyyy hh:mm"
        ws.cell(row=row, column=3, value=order["customer"]["name"])
        ws.cell(row=row, column=4, value=order["customer"].get("phone", ""))
        ws.cell(row=row, column=5,
                value=sum(item["quantity"] for item in order["items"]))
        for col, value in ((6, order["subtotal"]), (7, order["discount"]),
                           (8, order["total"]), (9, refunded)):
            cell = ws.cell(row=row, column=col, value=value)
            cell.number_format = MONEY_FMT
        ws.cell(row=row, column=10, value=status)

    # dong tong: cong don chua huy va TRU phan da hoan tra, dung cong thuc
    # de nguoi nhan file loc / sua trong Excel thi tong van tu tinh lai
    total_row = len(orders) + 2
    label = ws.cell(row=total_row, column=1,
                    value="Tổng thực thu (không tính đơn hủy, trừ hoàn trả)")
    label.font = Font(bold=True)
    if orders:
        last = total_row - 1
        formula = (f'=SUMIF(J2:J{last},"<>Đã hủy",H2:H{last})'
                   f'-SUMIF(J2:J{last},"<>Đã hủy",I2:I{last})')
        cell = ws.cell(row=total_row, column=8, value=formula)
        cell.font = Font(bold=True)
        cell.number_format = MONEY_FMT

    return _save(wb, "hoa_don")


def export_report(date_from=None, date_to=None, group_by: str = "month"):
    """Xuat bao cao thong ke 4 sheet, so lieu lay tu cung cac aggregation
    pipeline ma man hinh Thong ke dang dung."""
    wb = Workbook()

    # ---- sheet 1: tong quan ----
    ws = wb.active
    ws.title = "Tổng quan"
    _write_header(ws, ["Chỉ số", "Giá trị"], [26, 20])
    stats = report_service.summary(date_from, date_to)
    rows = [
        ("Doanh thu", stats["revenue"], MONEY_FMT),
        ("Giá vốn", stats["cost"], MONEY_FMT),
        ("Lợi nhuận", stats["profit"], MONEY_FMT),
        ("Số hóa đơn", stats["orders"], "#,##0"),
        ("Sản phẩm đã bán", stats["products_sold"], "#,##0"),
        ("Trung bình mỗi đơn", stats["avg_order"], MONEY_FMT),
    ]
    fmt_range = "(toàn bộ dữ liệu)"
    if date_from or date_to:
        start = f"{date_from:%d/%m/%Y}" if date_from else "..."
        end = f"{date_to:%d/%m/%Y}" if date_to else "..."
        fmt_range = f"từ {start} đến {end}"
    rows.append(("Khoảng thời gian", fmt_range, None))
    for index, (label, value, fmt) in enumerate(rows, start=2):
        ws.cell(row=index, column=1, value=label)
        cell = ws.cell(row=index, column=2, value=value)
        if fmt:
            cell.number_format = fmt

    # ---- sheet 2: theo thoi gian ----
    ws = wb.create_sheet("Theo thời gian")
    _write_header(ws, ["Kỳ", "Doanh thu", "Số đơn"], [14, 18, 10])
    for index, row in enumerate(
            report_service.revenue_by_period(group_by, date_from, date_to),
            start=2):
        ws.cell(row=index, column=1, value=row["_id"])
        ws.cell(row=index, column=2, value=row["revenue"]).number_format = MONEY_FMT
        ws.cell(row=index, column=3, value=row["orders"])

    # ---- sheet 3: theo danh muc ----
    ws = wb.create_sheet("Theo danh mục")
    _write_header(ws, ["Danh mục", "Doanh thu", "Số lượng"], [20, 18, 10])
    for index, row in enumerate(
            report_service.revenue_by_category(date_from, date_to), start=2):
        ws.cell(row=index, column=1, value=row["_id"])
        ws.cell(row=index, column=2, value=row["revenue"]).number_format = MONEY_FMT
        ws.cell(row=index, column=3, value=row["quantity"])

    # ---- sheet 4: top san pham ----
    ws = wb.create_sheet("Top sản phẩm")
    _write_header(ws, ["Sản phẩm", "Đã bán", "Doanh thu"], [34, 10, 18])
    for index, row in enumerate(
            report_service.top_products(10, date_from, date_to), start=2):
        ws.cell(row=index, column=1, value=row["_id"])
        ws.cell(row=index, column=2, value=row["quantity"])
        ws.cell(row=index, column=3, value=row["revenue"]).number_format = MONEY_FMT

    return _save(wb, "thong_ke")
